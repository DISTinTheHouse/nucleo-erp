from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.middleware.csrf import get_token
from django.views.decorators.csrf import csrf_protect

from .models import Producto


@login_required
@csrf_protect
def importar_productos_excel(request):
    if request.method == "GET":
        csrf_token = get_token(request)
        return HttpResponse(
            f"""
            <html>
              <body>
                <h2>Importar Productos (Excel)</h2>
                <p>Sube un archivo .xlsx con una columna llamada <b>nombre</b> o con el nombre en la primera columna.</p>
                <form method="post" enctype="multipart/form-data">
                  <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                  <input type="file" name="archivo" accept=".xlsx,.xlsm,.xltx,.xltm" required />
                  <button type="submit">Importar</button>
                </form>
              </body>
            </html>
            """
        )

    if request.method != "POST":
        return HttpResponseBadRequest("Método no permitido.")

    archivo = request.FILES.get("archivo")
    if not archivo:
        return HttpResponseBadRequest("Falta el archivo.")

    user = request.user
    empresa = getattr(user, "empresa", None)
    if not empresa:
        return HttpResponseForbidden("El usuario no tiene empresa asignada.")
    if not (getattr(user, "is_superuser", False) or getattr(user, "is_admin_empresa", False)):
        return HttpResponseForbidden("No autorizado.")

    try:
        from openpyxl import load_workbook
    except Exception:
        return HttpResponseBadRequest("Dependencia faltante para leer Excel (openpyxl).")

    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return HttpResponseBadRequest("No se pudo leer el archivo. Asegúrate que sea un .xlsx válido.")

    rows = ws.iter_rows(values_only=True)
    first_row = next(rows, None)
    if not first_row:
        return HttpResponseBadRequest("El archivo está vacío.")

    header = [
        str(v).strip().lower() if v is not None else ""
        for v in first_row
    ]
    if "nombre" in header:
        nombre_idx = header.index("nombre")
        has_header = True
    else:
        nombre_idx = 0
        has_header = False

    nombres = []

    def iter_data_rows():
        if not has_header:
            yield first_row
        for r in rows:
            yield r

    for r in iter_data_rows():
        if not r or len(r) <= nombre_idx:
            continue
        value = r[nombre_idx]
        if value is None:
            continue
        nombre = str(value).strip()
        if not nombre:
            continue
        nombres.append(nombre)

    if not nombres:
        return HttpResponseBadRequest("No se encontraron nombres de producto en el archivo.")

    nombres = list(dict.fromkeys(nombres))

    with transaction.atomic():
        existentes = set(
            Producto.objects.filter(empresa=empresa, nombre__in=nombres).values_list("nombre", flat=True)
        )
        nuevos = [
            Producto(empresa=empresa, nombre=n)
            for n in nombres
            if n not in existentes
        ]
        if nuevos:
            Producto.objects.bulk_create(nuevos)

    return HttpResponse(
        f"Importación completada. Leídos: {len(nombres)}. Creados: {len(nuevos)}. Ya existían: {len(existentes)}."
    )
